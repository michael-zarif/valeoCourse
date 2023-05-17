from openpyxl import load_workbook

wb = load_workbook('new_excel.xlsx')
print(wb.sheetnames)
sheet_1 = wb['Sheet1']
print(sheet_1['A1'].value)
sheet_1['A1'] = 'number'
print(sheet_1['A1'].value)

